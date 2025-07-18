#!/usr/bin/env python3
"""
Excel处理器 - Excel文件处理和分析工具

功能:
- Excel文件读写和格式转换
- 数据清洗和验证
- 图表生成和样式设置
- 多工作表处理
- 数据透视表生成
- 公式计算和验证

作者: ToolCollection Team
版本: 1.0.0
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart
import argparse
import sys
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import logging
from datetime import datetime
import json

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Excel处理器类"""
    
    def __init__(self, input_file: Optional[str] = None):
        """
        初始化Excel处理器
        
        Args:
            input_file: 输入Excel文件路径
        """
        self.input_file = Path(input_file) if input_file else None
        self.workbook = None
        self.dataframes = {}
    
    def load_excel(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, pd.DataFrame]:
        """
        加载Excel文件
        
        Args:
            file_path: Excel文件路径
            sheet_name: 指定工作表名称，None表示加载所有工作表
            
        Returns:
            工作表数据字典
        """
        try:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                self.dataframes[sheet_name] = df
                logger.info(f"成功加载工作表: {sheet_name}")
            else:
                self.dataframes = pd.read_excel(file_path, sheet_name=None)
                logger.info(f"成功加载Excel文件: {file_path}, 包含 {len(self.dataframes)} 个工作表")
            
            return self.dataframes
        except Exception as e:
            logger.error(f"加载Excel文件失败: {e}")
            raise
    
    def get_sheet_info(self) -> Dict[str, Any]:
        """
        获取工作表信息
        
        Returns:
            工作表信息
        """
        if not self.dataframes:
            return {}
        
        info = {}
        for sheet_name, df in self.dataframes.items():
            info[sheet_name] = {
                'rows': len(df),
                'columns': len(df.columns),
                'column_names': list(df.columns),
                'data_types': df.dtypes.to_dict(),
                'missing_values': df.isnull().sum().to_dict()
            }
        
        return info
    
    def clean_data(self, sheet_name: str, 
                   remove_duplicates: bool = True,
                   fill_missing: str = 'forward',
                   remove_empty_rows: bool = True) -> pd.DataFrame:
        """
        清洗数据
        
        Args:
            sheet_name: 工作表名称
            remove_duplicates: 是否删除重复行
            fill_missing: 填充缺失值的方法
            remove_empty_rows: 是否删除空行
            
        Returns:
            清洗后的数据框
        """
        if sheet_name not in self.dataframes:
            raise ValueError(f"工作表 {sheet_name} 不存在")
        
        df = self.dataframes[sheet_name].copy()
        
        # 删除重复行
        if remove_duplicates:
            original_rows = len(df)
            df = df.drop_duplicates()
            removed_rows = original_rows - len(df)
            logger.info(f"删除了 {removed_rows} 个重复行")
        
        # 填充缺失值
        if fill_missing == 'forward':
            df = df.fillna(method='ffill')
        elif fill_missing == 'backward':
            df = df.fillna(method='bfill')
        elif fill_missing == 'zero':
            df = df.fillna(0)
        elif fill_missing == 'mean':
            df = df.fillna(df.mean())
        
        # 删除空行
        if remove_empty_rows:
            original_rows = len(df)
            df = df.dropna(how='all')
            removed_rows = original_rows - len(df)
            logger.info(f"删除了 {removed_rows} 个空行")
        
        self.dataframes[sheet_name] = df
        return df
    
    def create_pivot_table(self, sheet_name: str, 
                          index: List[str], 
                          columns: Optional[List[str]] = None,
                          values: Optional[List[str]] = None,
                          aggfunc: str = 'mean') -> pd.DataFrame:
        """
        创建数据透视表
        
        Args:
            sheet_name: 工作表名称
            index: 行索引列
            columns: 列索引列
            values: 值列
            aggfunc: 聚合函数
            
        Returns:
            数据透视表
        """
        if sheet_name not in self.dataframes:
            raise ValueError(f"工作表 {sheet_name} 不存在")
        
        df = self.dataframes[sheet_name]
        
        pivot = pd.pivot_table(
            df, 
            index=index,
            columns=columns,
            values=values,
            aggfunc=aggfunc,
            fill_value=0
        )
        
        return pivot
    
    def add_formulas(self, sheet_name: str, 
                    formula_column: str, 
                    formula: str,
                    start_row: int = 2) -> None:
        """
        添加公式列
        
        Args:
            sheet_name: 工作表名称
            formula_column: 公式列名
            formula: 公式字符串
            start_row: 开始行号
        """
        if sheet_name not in self.dataframes:
            raise ValueError(f"工作表 {sheet_name} 不存在")
        
        df = self.dataframes[sheet_name]
        
        # 简单的公式计算
        if formula == 'sum':
            df[formula_column] = df.select_dtypes(include=[np.number]).sum(axis=1)
        elif formula == 'average':
            df[formula_column] = df.select_dtypes(include=[np.number]).mean(axis=1)
        elif formula == 'count':
            df[formula_column] = df.count(axis=1)
        else:
            # 自定义公式
            try:
                df[formula_column] = df.eval(formula)
            except Exception as e:
                logger.error(f"公式计算失败: {e}")
                raise
        
        self.dataframes[sheet_name] = df
        logger.info(f"已添加公式列: {formula_column}")
    
    def create_chart(self, sheet_name: str, 
                    chart_type: str,
                    x_column: str,
                    y_columns: List[str],
                    title: str = "",
                    output_file: Optional[str] = None) -> None:
        """
        创建图表
        
        Args:
            sheet_name: 工作表名称
            chart_type: 图表类型 (bar, line, pie)
            x_column: X轴列
            y_columns: Y轴列
            title: 图表标题
            output_file: 输出文件路径
        """
        if sheet_name not in self.dataframes:
            raise ValueError(f"工作表 {sheet_name} 不存在")
        
        df = self.dataframes[sheet_name]
        
        # 创建新的工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 写入数据
        for col_idx, col in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col)
        
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 创建图表
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'line':
            chart = LineChart()
        elif chart_type == 'pie':
            chart = PieChart()
        else:
            raise ValueError(f"不支持的图表类型: {chart_type}")
        
        chart.title = title
        chart.x_axis.title = x_column
        chart.y_axis.title = "值"
        
        # 添加数据
        data = openpyxl.chart.Reference(ws, min_col=2, min_row=1, max_row=len(df)+1, max_col=len(y_columns)+1)
        cats = openpyxl.chart.Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "E2")
        
        # 保存文件
        if output_file:
            wb.save(output_file)
            logger.info(f"图表已保存到: {output_file}")
        else:
            wb.save(f"chart_{sheet_name}_{chart_type}.xlsx")
            logger.info(f"图表已保存到: chart_{sheet_name}_{chart_type}.xlsx")
    
    def apply_styling(self, sheet_name: str,
                     header_style: bool = True,
                     alternate_rows: bool = True,
                     borders: bool = True) -> None:
        """
        应用样式
        
        Args:
            sheet_name: 工作表名称
            header_style: 是否设置表头样式
            alternate_rows: 是否设置交替行样式
            borders: 是否添加边框
        """
        if sheet_name not in self.dataframes:
            raise ValueError(f"工作表 {sheet_name} 不存在")
        
        df = self.dataframes[sheet_name]
        
        # 创建新的工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 写入数据
        for col_idx, col in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col)
        
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 设置表头样式
        if header_style:
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
        
        # 设置交替行样式
        if alternate_rows:
            alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for row_idx in range(2, len(df) + 2, 2):
                for cell in ws[row_idx]:
                    cell.fill = alt_fill
        
        # 添加边框
        if borders:
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
        
        # 保存文件
        wb.save(f"styled_{sheet_name}.xlsx")
        logger.info(f"样式化文件已保存到: styled_{sheet_name}.xlsx")
    
    def merge_sheets(self, output_sheet: str = "合并数据") -> pd.DataFrame:
        """
        合并所有工作表
        
        Args:
            output_sheet: 输出工作表名称
            
        Returns:
            合并后的数据框
        """
        if not self.dataframes:
            raise ValueError("没有数据可合并")
        
        merged_df = pd.concat(self.dataframes.values(), ignore_index=True)
        self.dataframes[output_sheet] = merged_df
        
        logger.info(f"已合并 {len(self.dataframes) - 1} 个工作表到 {output_sheet}")
        return merged_df
    
    def save_excel(self, output_file: str, 
                  sheet_name: Optional[str] = None,
                  include_index: bool = False) -> None:
        """
        保存Excel文件
        
        Args:
            output_file: 输出文件路径
            sheet_name: 指定工作表名称
            include_index: 是否包含索引
        """
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                if sheet_name:
                    if sheet_name in self.dataframes:
                        self.dataframes[sheet_name].to_excel(
                            writer, sheet_name=sheet_name, index=include_index
                        )
                else:
                    for name, df in self.dataframes.items():
                        df.to_excel(writer, sheet_name=name, index=include_index)
            
            logger.info(f"Excel文件已保存到: {output_file}")
        except Exception as e:
            logger.error(f"保存Excel文件失败: {e}")
            raise
    
    def validate_data(self, sheet_name: str, 
                     rules: Dict[str, Dict[str, Any]]) -> Dict[str, List[str]]:
        """
        数据验证
        
        Args:
            sheet_name: 工作表名称
            rules: 验证规则
            
        Returns:
            验证结果
        """
        if sheet_name not in self.dataframes:
            raise ValueError(f"工作表 {sheet_name} 不存在")
        
        df = self.dataframes[sheet_name]
        errors = {}
        
        for column, rule in rules.items():
            if column not in df.columns:
                errors[column] = [f"列 {column} 不存在"]
                continue
            
            column_errors = []
            
            # 数据类型验证
            if 'type' in rule:
                expected_type = rule['type']
                if expected_type == 'numeric':
                    if not pd.api.types.is_numeric_dtype(df[column]):
                        column_errors.append(f"列 {column} 应该为数值类型")
                elif expected_type == 'string':
                    if not pd.api.types.is_string_dtype(df[column]):
                        column_errors.append(f"列 {column} 应该为字符串类型")
                elif expected_type == 'date':
                    if not pd.api.types.is_datetime64_any_dtype(df[column]):
                        column_errors.append(f"列 {column} 应该为日期类型")
            
            # 范围验证
            if 'min' in rule:
                min_val = rule['min']
                invalid_count = (df[column] < min_val).sum()
                if invalid_count > 0:
                    column_errors.append(f"列 {column} 有 {invalid_count} 个值小于最小值 {min_val}")
            
            if 'max' in rule:
                max_val = rule['max']
                invalid_count = (df[column] > max_val).sum()
                if invalid_count > 0:
                    column_errors.append(f"列 {column} 有 {invalid_count} 个值大于最大值 {max_val}")
            
            # 唯一性验证
            if rule.get('unique', False):
                duplicate_count = df[column].duplicated().sum()
                if duplicate_count > 0:
                    column_errors.append(f"列 {column} 有 {duplicate_count} 个重复值")
            
            # 非空验证
            if rule.get('required', False):
                null_count = df[column].isnull().sum()
                if null_count > 0:
                    column_errors.append(f"列 {column} 有 {null_count} 个空值")
            
            if column_errors:
                errors[column] = column_errors
        
        return errors


def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='Excel文件处理和分析工具')
    parser.add_argument('input', help='输入Excel文件路径')
    parser.add_argument('-o', '--output', help='输出文件路径')
    parser.add_argument('-s', '--sheet', help='指定工作表名称')
    parser.add_argument('--info', action='store_true', help='显示工作表信息')
    parser.add_argument('--clean', action='store_true', help='清洗数据')
    parser.add_argument('--pivot', nargs='+', help='创建数据透视表 (index columns values)')
    parser.add_argument('--chart', nargs='+', help='创建图表 (type x_column y_columns...)')
    parser.add_argument('--style', action='store_true', help='应用样式')
    parser.add_argument('--merge', action='store_true', help='合并所有工作表')
    parser.add_argument('--validate', help='验证规则JSON文件路径')
    
    args = parser.parse_args()
    
    try:
        # 创建处理器
        processor = ExcelProcessor()
        
        # 加载Excel文件
        processor.load_excel(args.input, args.sheet)
        
        # 显示工作表信息
        if args.info:
            info = processor.get_sheet_info()
            print("工作表信息:")
            for sheet_name, sheet_info in info.items():
                print(f"\n{sheet_name}:")
                print(f"  行数: {sheet_info['rows']}")
                print(f"  列数: {sheet_info['columns']}")
                print(f"  列名: {sheet_info['column_names']}")
                print(f"  缺失值: {sheet_info['missing_values']}")
        
        # 清洗数据
        if args.clean:
            if args.sheet:
                processor.clean_data(args.sheet)
            else:
                for sheet_name in processor.dataframes.keys():
                    processor.clean_data(sheet_name)
        
        # 创建数据透视表
        if args.pivot and len(args.pivot) >= 3:
            index = args.pivot[0].split(',')
            columns = args.pivot[1].split(',') if args.pivot[1] != 'None' else None
            values = args.pivot[2].split(',') if args.pivot[2] != 'None' else None
            
            sheet_name = args.sheet or list(processor.dataframes.keys())[0]
            pivot = processor.create_pivot_table(sheet_name, index, columns, values)
            print("数据透视表:")
            print(pivot)
        
        # 创建图表
        if args.chart and len(args.chart) >= 3:
            chart_type = args.chart[0]
            x_column = args.chart[1]
            y_columns = args.chart[2].split(',')
            
            sheet_name = args.sheet or list(processor.dataframes.keys())[0]
            processor.create_chart(sheet_name, chart_type, x_column, y_columns)
        
        # 应用样式
        if args.style:
            if args.sheet:
                processor.apply_styling(args.sheet)
            else:
                for sheet_name in processor.dataframes.keys():
                    processor.apply_styling(sheet_name)
        
        # 合并工作表
        if args.merge:
            merged_df = processor.merge_sheets()
            print(f"合并后的数据形状: {merged_df.shape}")
        
        # 数据验证
        if args.validate:
            with open(args.validate, 'r', encoding='utf-8') as f:
                rules = json.load(f)
            
            if args.sheet:
                errors = processor.validate_data(args.sheet, rules)
            else:
                errors = {}
                for sheet_name in processor.dataframes.keys():
                    sheet_errors = processor.validate_data(sheet_name, rules)
                    if sheet_errors:
                        errors[sheet_name] = sheet_errors
            
            if errors:
                print("数据验证错误:")
                for sheet_or_col, error_list in errors.items():
                    print(f"\n{sheet_or_col}:")
                    for error in error_list:
                        print(f"  - {error}")
            else:
                print("数据验证通过")
        
        # 保存文件
        if args.output:
            processor.save_excel(args.output, args.sheet)
        
    except Exception as e:
        logger.error(f"处理失败: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main() 